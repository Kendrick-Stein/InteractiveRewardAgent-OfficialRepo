from __future__ import annotations

import json
import os
import re
import shlex
import subprocess
import sys
import tempfile
import time
from dataclasses import dataclass
from typing import Any, Dict, Optional, Tuple

# smolagents Tool base
from smolagents import Tool

# DeerAPI (OpenAI-compatible)
API_URL = (os.getenv("IRA_BASE_URL") or os.getenv("OPENAI_BASE_URL") or "https://api.cometapi.com/v1/").rstrip("/") + "/"
MODEL_NAME = "o3"

# Where to place the temporary generated checker and logs
GENERATED_CHECKER_PATH = os.path.join(tempfile.gettempdir(), f"ira_excel_checker_{os.getpid()}.py")
HTTP_RETRY = 3
HTTP_TIMEOUT_SEC = 120


# ==== Auto-injected extraction helpers for XLSX (from excel-test/exame.py, trimmed) ====
EXTRACT_PRELUDE = r"""
# ==== Auto-injected extraction helpers for XLSX (simplified from extractexcelinfo.py) ====
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


def color_to_hex(color) -> Optional[str]:
    try:
        if color is None:
            return None
        rgb = getattr(color, 'rgb', None)
        if rgb:
            return str(rgb)
        tint = getattr(color, 'tint', None)
        typ = getattr(color, 'type', None)
        return f"{typ}:{tint}" if typ is not None else None
    except Exception:
        return None


def font_summary(font) -> Dict[str, Any]:
    try:
        return {
            'name': getattr(font, 'name', None),
            'size': getattr(font, 'size', None),
            'bold': getattr(font, 'bold', None),
            'italic': getattr(font, 'italic', None),
            'underline': getattr(font, 'underline', None),
            'color': color_to_hex(getattr(font, 'color', None)),
        }
    except Exception:
        return {}


def fill_summary(fill) -> Dict[str, Any]:
    try:
        return {
            'patternType': getattr(fill, 'patternType', None),
            'fgColor': color_to_hex(getattr(fill, 'fgColor', None)),
            'bgColor': color_to_hex(getattr(fill, 'bgColor', None)),
        }
    except Exception:
        return {}


def border_side_summary(side) -> Dict[str, Any]:
    try:
        return {
            'style': getattr(side, 'style', None),
            'color': color_to_hex(getattr(side, 'color', None)),
        }
    except Exception:
        return {}


def border_summary(border) -> Dict[str, Any]:
    try:
        return {
            'left': border_side_summary(getattr(border, 'left', None)),
            'right': border_side_summary(getattr(border, 'right', None)),
            'top': border_side_summary(getattr(border, 'top', None)),
            'bottom': border_side_summary(getattr(border, 'bottom', None)),
        }
    except Exception:
        return {}


def alignment_summary(alignment) -> Dict[str, Any]:
    try:
        return {
            'horizontal': getattr(alignment, 'horizontal', None),
            'vertical': getattr(alignment, 'vertical', None),
            'wrapText': getattr(alignment, 'wrapText', None),
            'shrinkToFit': getattr(alignment, 'shrinkToFit', None),
            'textRotation': getattr(alignment, 'textRotation', None),
            'indent': getattr(alignment, 'indent', None),
        }
    except Exception:
        return {}


def protection_summary(protection) -> Dict[str, Any]:
    try:
        return {
            'locked': getattr(protection, 'locked', None),
            'hidden': getattr(protection, 'hidden', None),
        }
    except Exception:
        return {}


def style_summary(cell) -> Dict[str, Any]:
    try:
        return {
            'number_format': getattr(cell, 'number_format', None),
            'font': font_summary(getattr(cell, 'font', None)),
            'fill': fill_summary(getattr(cell, 'fill', None)),
            'border': border_summary(getattr(cell, 'border', None)),
            'alignment': alignment_summary(getattr(cell, 'alignment', None)),
            'protection': protection_summary(getattr(cell, 'protection', None)),
        }
    except Exception:
        return {}


def has_meaningful_style(st: Dict[str, Any]) -> bool:
    try:
        if not st:
            return False
        nf = st.get('number_format')
        if nf and nf not in ('General', 'general'):
            return True
        font = st.get('font') or {}
        if any(font.get(k) for k in ('bold', 'italic', 'underline', 'color', 'name')):
            return True
        fill = st.get('fill') or {}
        if any(fill.get(k) for k in ('patternType', 'fgColor', 'bgColor')):
            return True
        border = st.get('border') or {}
        for side in ('left', 'right', 'top', 'bottom'):
            if (border.get(side) or {}).get('style'):
                return True
        align = st.get('alignment') or {}
        if any(align.get(k) for k in ('horizontal', 'vertical', 'wrapText', 'shrinkToFit', 'textRotation', 'indent')):
            return True
        prot = st.get('protection') or {}
        if any(prot.get(k) for k in ('locked', 'hidden')):
            return True
    except Exception:
        pass
    return False


def cell_is_meaningful(cell, cached_value) -> bool:
    try:
        if cell.value is not None:
            return True
        if getattr(cell, 'data_type', None) == 'f':
            return True
        if getattr(cell, 'hyperlink', None):
            return True
        st = style_summary(cell)
        if has_meaningful_style(st):
            return True
        if cached_value is not None:
            return True
    except Exception:
        pass
    return False


def extract_cells(ws, ws_cached=None):
    cells = []
    try:
        dim = ws.calculate_dimension()
        min_col, min_row, max_col, max_row = range_boundaries(dim)
        cached = {}
        try:
            if ws_cached is not None:
                for row in ws_cached.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    for c in row:
                        cached[c.coordinate] = c.value
        except Exception:
            cached = {}
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cached_value = cached.get(cell.coordinate)
                if not cell_is_meaningful(cell, cached_value):
                    continue
                st = style_summary(cell)
                info = {
                    'coord': cell.coordinate,
                    'value': cell.value,
                    'data_type': getattr(cell, 'data_type', None),
                    'formula': None,
                    'cached_value': cached_value,
                    'number_format': st.get('number_format'),
                    'style': st,
                    'hyperlink': None,
                }
                try:
                    if getattr(cell, 'data_type', None) == 'f':
                        info['formula'] = cell.value
                except Exception:
                    pass
                try:
                    hl = getattr(cell, 'hyperlink', None)
                    if hl:
                        info['hyperlink'] = {
                            'address': getattr(hl, 'target', None),
                            'location': getattr(hl, 'location', None),
                            'display': getattr(hl, 'display', None),
                        }
                except Exception:
                    pass
                cells.append(info)
    except Exception:
        pass
    return cells


def merged_ranges(ws):
    out = []
    try:
        for rng in getattr(ws, 'merged_cells', []).ranges:
            try:
                out.append(str(rng))
            except Exception:
                continue
    except Exception:
        pass
    return out


def tables_summary(ws):
    out = []
    try:
        tbls = getattr(ws, 'tables', {})
        for name, tbl in tbls.items():
            out.append({
                'name': getattr(tbl, 'name', None) or name,
                'displayName': getattr(tbl, 'displayName', None),
                'ref': getattr(tbl, 'ref', None),
            })
    except Exception:
        pass
    return out


def data_validations_summary(ws):
    out = []
    try:
        dvs = getattr(ws, 'data_validations', None)
        if not dvs:
            return out
        for dv in getattr(dvs, 'dataValidation', []):
            item = {
                'type': getattr(dv, 'type', None),
                'operator': getattr(dv, 'operator', None),
                'allowBlank': getattr(dv, 'allowBlank', None),
                'showErrorMessage': getattr(dv, 'showErrorMessage', None),
                'showInputMessage': getattr(dv, 'showInputMessage', None),
                'promptTitle': getattr(dv, 'promptTitle', None),
                'prompt': getattr(dv, 'prompt', None),
                'errorTitle': getattr(dv, 'errorTitle', None),
                'error': getattr(dv, 'error', None),
                'formula1': getattr(dv, 'formula1', None),
                'formula2': getattr(dv, 'formula2', None),
                'sqref': getattr(dv, 'sqref', None),
            }
            out.append(item)
    except Exception:
        pass
    return out


def hyperlinks_summary(ws):
    out = []
    try:
        dim = ws.calculate_dimension()
        min_col, min_row, max_col, max_row = range_boundaries(dim)
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                hl = getattr(cell, 'hyperlink', None)
                if hl:
                    try:
                        out.append({
                            'cell': cell.coordinate,
                            'address': getattr(hl, 'target', None),
                            'location': getattr(hl, 'location', None),
                            'display': getattr(hl, 'display', None),
                        })
                    except Exception:
                        continue
    except Exception:
        pass
    return out


def charts_summary(ws):
    out = []
    try:
        charts = getattr(ws, '_charts', [])
        for ch in charts:
            item = {
                'type': ch.__class__.__name__,
                'anchor': getattr(ch, 'anchor', None),
                'title': None,
            }
            try:
                t = getattr(ch, 'title', None)
                item['title'] = str(t) if t is not None else None
            except Exception:
                pass
            out.append(item)
    except Exception:
        pass
    return out


def images_summary(ws):
    out = []
    try:
        imgs = getattr(ws, '_images', [])
        for im in imgs:
            item = {
                'type': im.__class__.__name__,
                'anchor': getattr(im, 'anchor', None),
                'width': getattr(im, 'width', None),
                'height': getattr(im, 'height', None),
                'format': getattr(im, 'format', None),
            }
            out.append(item)
    except Exception:
        pass
    return out


def sheet_view_summary(ws):
    out = {}
    try:
        sv = getattr(ws, 'sheet_view', None)
        if sv is not None:
            out = {
                'showGridLines': getattr(sv, 'showGridLines', None),
                'rightToLeft': getattr(sv, 'rightToLeft', None),
                'zoomScale': getattr(sv, 'zoomScale', None),
                'zoomScaleNormal': getattr(sv, 'zoomScaleNormal', None),
                'showRowColHeaders': getattr(sv, 'showRowColHeaders', None),
            }
    except Exception:
        pass
    return out


def page_setup_summary(ws):
    out = {}
    try:
        ps = getattr(ws, 'page_setup', None)
        pm = getattr(ws, 'page_margins', None)
        hf = getattr(ws, 'header_footer', None)
        out['page_setup'] = {
            'orientation': getattr(ps, 'orientation', None) if ps else None,
            'paperSize': getattr(ps, 'paperSize', None) if ps else None,
            'scale': getattr(ps, 'scale', None) if ps else None,
            'fitToHeight': getattr(ps, 'fitToHeight', None) if ps else None,
            'fitToWidth': getattr(ps, 'fitToWidth', None) if ps else None,
        }
        out['page_margins'] = {
            'left': getattr(pm, 'left', None) if pm else None,
            'right': getattr(pm, 'right', None) if pm else None,
            'top': getattr(pm, 'top', None) if pm else None,
            'bottom': getattr(pm, 'bottom', None) if pm else None,
            'header': getattr(pm, 'header', None) if pm else None,
            'footer': getattr(pm, 'footer', None) if pm else None,
        }
        out['header_footer'] = {
            'left_header': getattr(hf, 'leftHeader', None) if hf else None,
            'center_header': getattr(hf, 'centerHeader', None) if hf else None,
            'right_header': getattr(hf, 'rightHeader', None) if hf else None,
            'left_footer': getattr(hf, 'leftFooter', None) if hf else None,
            'center_footer': getattr(hf, 'centerFooter', None) if hf else None,
            'right_footer': getattr(hf, 'rightFooter', None) if hf else None,
        }
    except Exception:
        pass
    return out


def protection_summary_ws(ws):
    out = {}
    try:
        prot = getattr(ws, 'protection', None)
        if prot is not None:
            out = {
                'sheet': getattr(prot, 'sheet', None),
                'password': getattr(prot, 'password', None),
                'formatCells': getattr(prot, 'formatCells', None),
                'formatColumns': getattr(prot, 'formatColumns', None),
                'formatRows': getattr(prot, 'formatRows', None),
                'insertColumns': getattr(prot, 'insertColumns', None),
                'insertRows': getattr(prot, 'insertRows', None),
                'selectLockedCells': getattr(prot, 'selectLockedCells', None),
                'selectUnlockedCells': getattr(prot, 'selectUnlockedCells', None),
            }
    except Exception:
        pass
    return out


def extract_sheet(ws, index: int, ws_cached=None):
    info = {
        'title': getattr(ws, 'title', None),
        'index': index,
        'state': getattr(ws, 'sheet_state', None),
        'dimension': None,
        'freeze_panes': getattr(ws, 'freeze_panes', None),
        'sheet_view': sheet_view_summary(ws),
        'page_setup': page_setup_summary(ws),
        'protection': protection_summary_ws(ws),
        'auto_filter': None,
        'merged_cells': merged_ranges(ws),
        'tables': tables_summary(ws),
        'data_validations': data_validations_summary(ws),
        'hyperlinks': hyperlinks_summary(ws),
        'charts': charts_summary(ws),
        'images': images_summary(ws),
        'cells': [],
    }
    try:
        info['dimension'] = ws.calculate_dimension()
    except Exception:
        info['dimension'] = None
    try:
        af = getattr(ws, 'auto_filter', None)
        info['auto_filter'] = getattr(af, 'ref', None) if af else None
    except Exception:
        info['auto_filter'] = None
    info['cells'] = extract_cells(ws, ws_cached)
    return info


def extract_defined_names(wb):
    out = []
    try:
        dn = getattr(wb, 'defined_names', None)
        if not dn:
            return out
        for defn in dn.definedName:
            item = {
                'name': getattr(defn, 'name', None),
                'comment': getattr(defn, 'comment', None),
                'localSheetId': getattr(defn, 'localSheetId', None),
                'hidden': getattr(defn, 'hidden', None),
                'destinations': [],
            }
            try:
                for sheetname, ref in dn.destinations(defn):
                    item['destinations'].append({'sheet': sheetname, 'ref': ref})
            except Exception:
                pass
            out.append(item)
    except Exception:
        pass
    return out


def extract_workbook_props(wb):
    info = {}
    try:
        props = getattr(wb, 'properties', None)
        if props is not None:
            info['properties'] = {
                'title': getattr(props, 'title', None),
                'subject': getattr(props, 'subject', None),
                'creator': getattr(props, 'creator', None),
                'lastModifiedBy': getattr(props, 'lastModifiedBy', None),
                'created': getattr(props, 'created', None),
                'modified': getattr(props, 'modified', None),
                'category': getattr(props, 'category', None),
                'description': getattr(props, 'description', None),
                'keywords': getattr(props, 'keywords', None),
            }
    except Exception:
        pass
    try:
        calc = getattr(wb, 'calculation_properties', None)
        if calc is not None:
            info['calculation'] = {
                'calcId': getattr(calc, 'calcId', None),
                'calcMode': getattr(calc, 'calcMode', None),
                'fullCalcOnLoad': getattr(calc, 'fullCalcOnLoad', None),
                'refMode': getattr(calc, 'refMode', None),
            }
    except Exception:
        pass
    try:
        theme = getattr(wb, 'loaded_theme', None)
        info['theme'] = {
            'exists': theme is not None,
            'length': len(theme) if theme is not None else None,
        }
    except Exception:
        pass
    info['defined_names'] = extract_defined_names(wb)
    return info


def extract_workbook(wb, wb_cached=None, file_path: str = None):
    out = {
        'workbook': {},
        'sheets': [],
    }
    out['workbook'] = extract_workbook_props(wb)
    if file_path:
        out['workbook']['file_path'] = file_path
    try:
        for idx, ws in enumerate(wb.worksheets):
            ws_cached = None
            try:
                if wb_cached is not None:
                    ws_cached = wb_cached[ws.title]
            except Exception:
                ws_cached = None
            out['sheets'].append(extract_sheet(ws, idx, ws_cached))
    except Exception:
        pass
    return out
# ==== End of helpers ====
"""


@dataclass
class VerifierResult:
    passed: bool
    reason: str
    llm_code_path: str
    stdout: str
    stderr: str

    def to_dict(self) -> Dict[str, Any]:
        return {
            "passed": self.passed,
            "reason": self.reason,
            "llm_code_path": self.llm_code_path,
            "stdout": self.stdout,
            "stderr": self.stderr,
        }


# ============ Internal helpers (adapted from excel-test/exame.py) ============

def _load_api_key() -> str:
    key = os.getenv("IRA_API_KEY") or os.getenv("OPENAI_API_KEY") or os.getenv("deerapi_key")
    if not key:
        try:
            from dotenv import load_dotenv
            load_dotenv()
            key = os.getenv("IRA_API_KEY") or os.getenv("OPENAI_API_KEY") or os.getenv("deerapi_key")
        except Exception:
            pass
    if not key:
        raise RuntimeError("Missing API key. Set IRA_API_KEY in the environment or a .env file.")
    return key


def _build_client() -> Any:
    api_key = _load_api_key()
    import openai  # type: ignore
    client = openai.OpenAI(api_key=api_key, base_url=API_URL)
    return client


def _call_llm_via_client(system: str, user: str) -> str:
    client = _build_client()
    resp = client.chat.completions.create(
        model=MODEL_NAME,
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
        temperature=0,
    )
    content = getattr(resp.choices[0].message, "content", None)
    if isinstance(content, str):
        return content
    try:
        return resp.choices[0].message["content"]  # type: ignore[index]
    except Exception:
        return ""


def _call_llm_via_http(system: str, user: str) -> str:
    api_key = _load_api_key()
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}",
    }
    payload = {
        "model": MODEL_NAME,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
        "temperature": 0,
    }
    last_err = None
    url = f"{API_URL}chat/completions"
    for i in range(1, HTTP_RETRY + 1):
        try:
            try:
                import requests  # type: ignore
                r = requests.post(url, headers=headers, json=payload, timeout=HTTP_TIMEOUT_SEC)
                r.raise_for_status()
                data = r.json()
            except Exception as e_req:
                from urllib.request import Request, urlopen
                import json as _json
                try:
                    req = Request(url, data=_json.dumps(payload).encode("utf-8"), headers=headers, method="POST")
                    with urlopen(req, timeout=HTTP_TIMEOUT_SEC) as resp:
                        txt = resp.read().decode("utf-8")
                    data = _json.loads(txt)
                except Exception as e_url:
                    last_err = e_url if e_req is None else e_req
                    raise last_err
            try:
                return data["choices"][0]["message"]["content"]
            except Exception as e:
                last_err = e
        except Exception as e:
            last_err = e
        if i < HTTP_RETRY:
            time.sleep(2 * i)
    if last_err:
        raise last_err
    return ""


def _llm_prompt(checks_text: str, error_context: Optional[str] = None) -> Tuple[str, str]:
    system = (
        "You are a careful Python auditor that writes a single, self-contained Python script "
        "to verify whether a set of simple checks are satisfied on an XLSX file. "
        "Output only one Python code block fenced by triple backticks. Do not add any extra text. "
        "Your script must:")
    system += (
        "\n- Use openpyxl to read the .xlsx file in read-only manner (no writes)."
        "\n- Implement: def check(xlsx_path: str, checks: str) -> dict with keys: passed (bool), details (str). The 'checks' is a simple string of atomic items (e.g., newline or semicolon separated) that the script should evaluate."
        "\n- Provide a __main__ that parses two required CLI args: --xlsx and --checks; then prints exactly one JSON line to stdout (no extra text)."
        "\n- Never modify files. No network. No os/subprocess/requests/shutil/socket/urllib usage. No eval/exec. Avoid sys.exit unless necessary."
        "\n- Choose which parts to inspect strictly from the provided checks text. Treat the checks as atomic items and evaluate them conservatively. If the checks don't specify a specific sheet/cell, you may check relevant areas and state your assumption in details."
        "\n- Be robust to values vs formulas; if the task mentions formulas, consider reading cached values by opening a second workbook with data_only=True."
        "\n- Inspect formatting when relevant: number_format, font, fill, border, alignment, merged cells, freeze panes, auto filter, data validation, hyperlinks, charts/images, etc."
        "\n- Your script must never crash; wrap main execution in try/except and always print one JSON line with keys 'passed' and 'details'."
    )

    system += (
        "\n\nNote: Predefined helper functions for robust XLSX extraction are injected at the top of the script and available to call: "
        "extract_workbook, extract_sheet, extract_cells, style_summary, color_to_hex, font_summary, fill_summary, border_summary, alignment_summary, protection_summary. "
        "Use these helpers to focus on the task-specific verification logic.")

    system += (
        "\n\nopenpyxl Quick Reference:" \
        "\n- Imports:" \
        "\n  from openpyxl import load_workbook" \
        "\n- Open workbook:" \
        "\n  wb = load_workbook(xlsx_path, data_only=False, read_only=False)" \
        "\n  wb_cached = load_workbook(xlsx_path, data_only=True, read_only=False)  # optional for cached values" \
        "\n- Iterate sheets & cells:" \
        "\n  for ws in wb.worksheets: ws.title; ws.calculate_dimension(); ws.iter_rows(...); cell.value; cell.data_type; cell.hyperlink; cell.number_format" \
        "\n- Other sheet properties:" \
        "\n  ws.freeze_panes; ws.auto_filter; ws.tables; ws.data_validations; ws.merged_cells; ws._charts; ws._images" \
        "\n- Output & safety:" \
        "\n  Print exactly one JSON line to stdout with keys 'passed' and 'details'. No file writes, no network, no os/subprocess/requests/shutil/socket/urllib, no eval/exec."
    )

    user = (
        "Write the script now. It must be a single code block:"\
        "\n```python\n# your script here\n```\n\n"\
        "The checks to verify are:\n" + checks_text + "\n"\
        "The script will be executed as: python excel_checker.py --xlsx /abs/path.xlsx --checks '<CHECKS>'\n"\
        "Remember to print only a single JSON line on stdout with keys: passed, details."
    )
    if error_context:
        user += ("\n\nThe previous attempt failed with the following error/output. "
                 "Regenerate a corrected script that avoids these issues and adheres to the rules.\n" + error_context)
    return system, user


def _extract_code_block(text: str) -> Optional[str]:
    pattern = r"```(?:python)?\n([\s\S]*?)\n```"
    m = re.search(pattern, text, re.IGNORECASE)
    if not m:
        return None
    return m.group(1)


def _basic_static_safety_check(code: str) -> Optional[str]:
    """Return None if safe enough, else reason."""
    forbidden_imports = [
        "os", "subprocess", "shutil", "socket", "requests", "urllib", "docx", "pptx", "pandas", "xlrd"
    ]
    for name in forbidden_imports:
        if re.search(rf"^\s*(import|from)\s+{re.escape(name)}\b", code, re.MULTILINE):
            return f"Forbidden import detected: {name}"
    if re.search(r"\beval\s*\(", code):
        return "Forbidden call: eval()"
    if re.search(r"\bexec\s*\(", code):
        return "Forbidden call: exec()"
    if re.search(r"open\s*\(.*['\"]\s*[wax]\s*['\"]", code):
        return "Forbidden file write mode detected"
    if re.search(r"http[s]?://", code):
        return "Forbidden network usage detected"
    return None


def _write_generated_checker(code: str, target_path: Optional[str] = None) -> str:
    path = target_path or GENERATED_CHECKER_PATH
    os.makedirs(os.path.dirname(path), exist_ok=True)
    prelude = EXTRACT_PRELUDE.strip() + "\n\n"
    with open(path, "w", encoding="utf-8") as f:
        f.write(prelude + code)
    return path


def _build_checker_cmd(checker_path: str, xlsx_path: str, checks_text: str) -> str:
    return f"{shlex.quote(sys.executable)} {shlex.quote(checker_path)} --xlsx {shlex.quote(xlsx_path)} --checks {shlex.quote(checks_text)}"


def _run_generated_checker(checker_path: str, xlsx_path: str, checks_text: str, timeout_sec: int = 20) -> Tuple[str, str, int, str]:
    cmd = _build_checker_cmd(checker_path, xlsx_path, checks_text)
    proc = subprocess.run(
        cmd,
        shell=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        timeout=timeout_sec,
        text=True,
    )
    return proc.stdout.strip(), proc.stderr.strip(), proc.returncode, cmd


def _parse_json_line(s: str) -> Optional[Dict[str, Any]]:
    line = s.strip().splitlines()[0] if s.strip() else ""
    candidate = line
    if not candidate or not candidate.strip().startswith("{"):
        m = re.search(r"\{[\s\S]*\}", s)
        candidate = m.group(0) if m else ""
    if not candidate:
        return None
    try:
        return json.loads(candidate)
    except Exception:
        return None


def _ensure_dir(path: str) -> None:
    try:
        os.makedirs(path, exist_ok=True)
    except Exception:
        pass


def _safe_write(path: str, content: str) -> None:
    try:
        _ensure_dir(os.path.dirname(path))
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
    except Exception:
        pass


def _log(trace: bool, msg: str) -> None:
    if trace:
        try:
            sys.stderr.write(msg.rstrip("\n") + "\n")
            sys.stderr.flush()
        except Exception:
            pass


def run_o3_verifier_excel(checks_text: str, xlsx_path: str, attempts: int = 3, trace: bool = False, log_root_dir: Optional[str] = None, output_dir: Optional[str] = None) -> VerifierResult:
    if not checks_text:
        raise ValueError("checks_text is required")
    if not xlsx_path:
        raise ValueError("xlsx_path is required")
    if not os.path.isabs(xlsx_path):
        xlsx_path = os.path.abspath(xlsx_path)
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"XLSX file not found: {xlsx_path}")

    ts = time.strftime("%Y%m%d-%H%M%S")
    if output_dir:
        base = os.path.join(output_dir, "doc_tools")
        log_root_dir = log_root_dir or os.path.join(base, "attempt_logs", f"excel-{ts}")
        checker_path = os.path.join(base, "excel_checker.py")
    else:
        if log_root_dir is None:
            log_root_dir = os.path.join(tempfile.gettempdir(), "ira_attempt_logs", f"excel-{ts}")
        checker_path = GENERATED_CHECKER_PATH
    _ensure_dir(log_root_dir)

    summary_path = os.path.join(log_root_dir, "summary.txt")

    attempt = 1
    error_ctx: Optional[str] = None
    last_stdout = ""
    last_stderr = ""
    last_code_path = ""
    last_content = ""

    while attempt <= max(1, attempts):
        attempt_dir = os.path.join(log_root_dir, f"attempt-{attempt}")
        _ensure_dir(attempt_dir)
        _log(trace, f"[excel attempt {attempt}] starting...")

        system, user = _llm_prompt(checks_text, error_context=error_ctx)
        _safe_write(os.path.join(attempt_dir, "00-prompt-system.txt"), system)
        _safe_write(os.path.join(attempt_dir, "01-prompt-user.txt"), user)

        content = ""
        err_sdk: Optional[Exception] = None
        used_path = "sdk"
        try:
            content = _call_llm_via_client(system, user)
        except Exception as e:
            err_sdk = e

        if not content or not _extract_code_block(content):
            used_path = "http"
            try:
                content = _call_llm_via_http(system, user)
            except Exception as e:
                meta = [
                    f"used_path={used_path}",
                    f"sdk_error={repr(err_sdk)}",
                    f"http_error={repr(e)}",
                ]
                _safe_write(os.path.join(attempt_dir, "02b-llm-meta.txt"), "\n".join(meta))
                reason = f"LLM call failed (sdk={err_sdk}) (http={e})"
                if attempt >= max(1, attempts):
                    _safe_write(summary_path, (open(summary_path, "r", encoding="utf-8").read() + "\n") if os.path.exists(summary_path) else "")
                    _safe_write(summary_path, f"attempt {attempt}: LLM call failed. {reason}\n")
                    _log(trace, f"[excel attempt {attempt}] LLM call failed. {reason}")
                    return VerifierResult(
                        passed=False,
                        reason=reason,
                        llm_code_path=last_code_path,
                        stdout=str(last_stdout or content),
                        stderr=str(last_stderr),
                    )
                else:
                    error_ctx = f"Previous LLM call failed. Reason: {reason}"
                    _safe_write(os.path.join(attempt_dir, "02-llm-raw.txt"), str(content))
                    _safe_write(os.path.join(attempt_dir, "02b-llm-meta.txt"), "\n".join(meta))
                    _safe_write(summary_path, (open(summary_path, "r", encoding="utf-8").read() + "\n") if os.path.exists(summary_path) else "")
                    _safe_write(summary_path, f"attempt {attempt}: LLM call failed. Will retry.\n")
                    _log(trace, f"[excel attempt {attempt}] LLM call failed, retrying...")
                    attempt += 1
                    continue

        last_content = content
        _safe_write(os.path.join(attempt_dir, "02-llm-raw.txt"), str(content))
        _safe_write(os.path.join(attempt_dir, "02b-llm-meta.txt"), f"used_path={used_path}\nsdk_error={repr(err_sdk)}")

        code = _extract_code_block(content)
        if not code:
            _safe_write(os.path.join(attempt_dir, "03-extracted_code.py"), code or "")
            if attempt >= max(1, attempts):
                _safe_write(summary_path, (open(summary_path, "r", encoding="utf-8").read() + "\n") if os.path.exists(summary_path) else "")
                _safe_write(summary_path, f"attempt {attempt}: no valid code block.\n")
                _log(trace, f"[excel attempt {attempt}] no valid code block; stop.")
                return VerifierResult(
                    passed=False,
                    reason="Failed to extract code block from LLM response",
                    llm_code_path=last_code_path,
                    stdout=content,
                    stderr="",
                )
            else:
                error_ctx = "The earlier response did not include a valid ```python code block. Ensure you wrap the entire script in triple backticks."
                _safe_write(summary_path, (open(summary_path, "r", encoding="utf-8").read() + "\n") if os.path.exists(summary_path) else "")
                _safe_write(summary_path, f"attempt {attempt}: no code block; retry.\n")
                _log(trace, f"[excel attempt {attempt}] no code block; retrying...")
                attempt += 1
                continue

        _safe_write(os.path.join(attempt_dir, "03-extracted_code.py"), code)

        unsafe_reason = _basic_static_safety_check(code)
        if unsafe_reason:
            _safe_write(os.path.join(attempt_dir, "04-safety_check.txt"), f"REJECTED: {unsafe_reason}")
            if attempt >= max(1, attempts):
                _safe_write(summary_path, (open(summary_path, "r", encoding="utf-8").read() + "\n") if os.path.exists(summary_path) else "")
                _safe_write(summary_path, f"attempt {attempt}: safety check failed: {unsafe_reason}.\n")
                _log(trace, f"[excel attempt {attempt}] safety check failed; stop.")
                return VerifierResult(
                    passed=False,
                    reason=f"Generated code rejected by safety check: {unsafe_reason}",
                    llm_code_path=last_code_path,
                    stdout=code,
                    stderr="",
                )
            else:
                error_ctx = f"Your previous code failed safety checks: {unsafe_reason}. Remove the offending imports/calls and regenerate."
                _safe_write(summary_path, (open(summary_path, "r", encoding="utf-8").read() + "\n") if os.path.exists(summary_path) else "")
                _safe_write(summary_path, f"attempt {attempt}: safety check failed; retry.\n")
                _log(trace, f"[excel attempt {attempt}] safety check failed; retrying...")
                attempt += 1
                continue
        else:
            _safe_write(os.path.join(attempt_dir, "04-safety_check.txt"), "OK")

        path = _write_generated_checker(code, target_path=checker_path)
        last_code_path = path
        per_attempt_copy = os.path.join(attempt_dir, f"generated_checker_attempt_{attempt}.py")
        _safe_write(per_attempt_copy, code)
        _safe_write(os.path.join(attempt_dir, "05-generated_checker_path.txt"), f"main={path}\ncopy={per_attempt_copy}")

        try:
            stdout, stderr, rc, cmd = _run_generated_checker(checker_path, xlsx_path, checks_text)
            _safe_write(os.path.join(attempt_dir, "06-run_command.txt"), cmd)
            _safe_write(os.path.join(attempt_dir, "07-stdout.txt"), stdout)
            _safe_write(os.path.join(attempt_dir, "08-stderr.txt"), stderr)
        except subprocess.TimeoutExpired:
            _safe_write(os.path.join(attempt_dir, "06-run_command.txt"), _build_checker_cmd(checker_path, xlsx_path, checks_text))
            _safe_write(os.path.join(attempt_dir, "08-stderr.txt"), "TimeoutExpired")
            if attempt >= max(1, attempts):
                _safe_write(summary_path, (open(summary_path, "r", encoding="utf-8").read() + "\n") if os.path.exists(summary_path) else "")
                _safe_write(summary_path, f"attempt {attempt}: checker timed out.\n")
                _log(trace, f"[excel attempt {attempt}] checker timed out; stop.")
                return VerifierResult(
                    passed=False,
                    reason="Generated checker timed out",
                    llm_code_path=path,
                    stdout=last_stdout,
                    stderr=last_stderr,
                )
            else:
                error_ctx = "Your previous script timed out. Make the logic efficient and avoid long loops."
                _safe_write(summary_path, (open(summary_path, "r", encoding="utf-8").read() + "\n") if os.path.exists(summary_path) else "")
                _safe_write(summary_path, f"attempt {attempt}: checker timed out; retry.\n")
                _log(trace, f"[excel attempt {attempt}] checker timed out; retrying...")
                attempt += 1
                continue
        except Exception as e:
            _safe_write(os.path.join(attempt_dir, "06-run_command.txt"), _build_checker_cmd(checker_path, xlsx_path, checks_text))
            _safe_write(os.path.join(attempt_dir, "08-stderr.txt"), f"Exception: {repr(e)}")
            if attempt >= max(1, attempts):
                _safe_write(summary_path, (open(summary_path, "r", encoding="utf-8").read() + "\n") if os.path.exists(summary_path) else "")
                _safe_write(summary_path, f"attempt {attempt}: failed to run checker: {repr(e)}.\n")
                _log(trace, f"[excel attempt {attempt}] failed to run checker; stop.")
                return VerifierResult(
                    passed=False,
                    reason=f"Failed to run generated checker: {e}",
                    llm_code_path=path,
                    stdout=last_stdout,
                    stderr=last_stderr,
                )
            else:
                error_ctx = f"Your previous script raised an exception on execution: {e}. Fix and regenerate."
                _safe_write(summary_path, (open(summary_path, "r", encoding="utf-8").read() + "\n") if os.path.exists(summary_path) else "")
                _safe_write(summary_path, f"attempt {attempt}: checker exception; retry.\n")
                _log(trace, f"[excel attempt {attempt}] checker exception; retrying...")
                attempt += 1
                continue

        last_stdout, last_stderr = stdout, stderr
        parsed = _parse_json_line(stdout)
        if parsed:
            _safe_write(os.path.join(attempt_dir, "09-parsed_json.json"), json.dumps(parsed, ensure_ascii=False, indent=2))
        else:
            _safe_write(os.path.join(attempt_dir, "09-parsed_json.json"), "<invalid or missing JSON line>")

        if not parsed or not isinstance(parsed, dict) or "passed" not in parsed:
            short_err = (stderr or "").strip()
            if len(short_err) > 800:
                short_err = short_err[-800:]
            if attempt >= max(1, attempts):
                _safe_write(summary_path, (open(summary_path, "r", encoding="utf-8").read() + "\n") if os.path.exists(summary_path) else "")
                _safe_write(summary_path, f"attempt {attempt}: invalid JSON output.\n")
                _log(trace, f"[excel attempt {attempt}] invalid JSON output; stop.")
                return VerifierResult(
                    passed=False,
                    reason="Generated checker did not produce a valid single-line JSON with 'passed'",
                    llm_code_path=path,
                    stdout=stdout,
                    stderr=stderr,
                )
            else:
                error_ctx = (
                    "Your script must print exactly one JSON line on stdout with keys 'passed' and 'details'. "
                    "The previous run failed. Here is stderr/issue summary:\n" + short_err + "\n\n"
                    "Also wrap your main in try/except so exceptions still result in a single JSON line."
                )
                _safe_write(summary_path, (open(summary_path, "r", encoding="utf-8").read() + "\n") if os.path.exists(summary_path) else "")
                _safe_write(summary_path, f"attempt {attempt}: invalid JSON output; retry.\n")
                _log(trace, f"[excel attempt {attempt}] invalid JSON; retrying...")
                attempt += 1
                continue

        passed = bool(parsed.get("passed"))
        details = str(parsed.get("details", ""))

        if not passed:
            short_details = details.strip()
            if len(short_details) > 800:
                short_details = short_details[:800]
            if attempt >= max(1, attempts):
                _safe_write(summary_path, (open(summary_path, "r", encoding="utf-8").read() + "\n") if os.path.exists(summary_path) else "")
                _safe_write(summary_path, f"attempt {attempt}: checker returned failed.\n")
                _log(trace, f"[excel attempt {attempt}] checker returned failed; stop.")
                return VerifierResult(
                    passed=False,
                    reason=details,
                    llm_code_path=path,
                    stdout=stdout,
                    stderr=stderr,
                )
            else:
                error_ctx = (
                    "Your previous checker returned passed=false with details: " + short_details + "\n"
                    "Regenerate a corrected checker that fixes this issue while strictly following all rules."
                )
                _safe_write(summary_path, (open(summary_path, "r", encoding="utf-8").read() + "\n") if os.path.exists(summary_path) else "")
                _safe_write(summary_path, f"attempt {attempt}: checker returned failed; retry.\n")
                _log(trace, f"[excel attempt {attempt}] checker returned failed; retrying...")
                last_stdout, last_stderr = stdout, stderr
                attempt += 1
                continue

        _safe_write(summary_path, (open(summary_path, "r", encoding="utf-8").read() + "\n") if os.path.exists(summary_path) else "")
        _safe_write(summary_path, f"attempt {attempt}: SUCCESS.\n")
        _log(trace, f"[excel attempt {attempt}] SUCCESS. Artifacts at: {attempt_dir}")
        return VerifierResult(
            passed=True,
            reason=details,
            llm_code_path=path,
            stdout=stdout,
            stderr=stderr,
        )

    _safe_write(summary_path, (open(summary_path, "r", encoding="utf-8").read() + "\n") if os.path.exists(summary_path) else "")
    _safe_write(summary_path, "No attempt yielded a valid result.\n")
    _log(trace, f"All attempts exhausted. Logs at: {log_root_dir}")
    return VerifierResult(
        passed=False,
        reason="Exhausted attempts without a valid result",
        llm_code_path=last_code_path,
        stdout=last_stdout or last_content,
        stderr=last_stderr,
    )


class CheckExcelFileTool(Tool):
    name = "checkexcelfile"
    description = (
        "Verify an Excel (.xlsx) file according to a simple checks string. "
        "This tool requires a Host file path. If your file is inside the VM (e.g., /home/user/...), "
        "please first use get_vm_file(vm_path, dest_name) to download it to the Host and then pass the returned Host path here. "
        "Returns a JSON string with keys passed, reason, llm_code_path, stdout, stderr."
    )
    inputs = {
        "file_path": {
            "description": (
                "Path to the .xlsx file on the Host (absolute or relative). "
                "Do not pass VM paths. If the file is in the VM, call get_vm_file first and then pass the returned Host path."
            ),
            "type": "string",
        },
        "things_to_check": {
            "description": "A simple string describing atomic checks to verify (e.g., newline/semicolon separated).",
            "type": "string",
        },
    }
    output_type = "string"

    def __init__(self, output_dir: Optional[str] = None):
        super().__init__()
        self.output_dir = output_dir

    def set_output_dir(self, output_dir: str) -> None:
        self.output_dir = output_dir

    def forward(self, file_path: str, things_to_check: str) -> str:
        return self.__call__(file_path, things_to_check)

    def __call__(self, file_path: str, things_to_check: str) -> str:
        try:
            # Early validate file path to avoid requiring API key when the file is missing
            if not file_path:
                return json.dumps({"passed": False, "reason": "file_path is required"}, ensure_ascii=False)
            if not things_to_check:
                return json.dumps({"passed": False, "reason": "things_to_check is required"}, ensure_ascii=False)
            abs_path = os.path.abspath(file_path)
            if not os.path.exists(abs_path):
                return json.dumps({
                    "passed": False,
                    "reason": (
                        f"File not found on host: {abs_path}. "
                        "This tool requires a Host file path. If your file is inside the VM (e.g., /home/user/...), "
                        "please use get_vm_file to download it to the Host and pass the returned Host path."
                    )
                }, ensure_ascii=False)

            res = run_o3_verifier_excel(things_to_check, abs_path, attempts=3, trace=False, log_root_dir=None, output_dir=self.output_dir)
            return json.dumps(res.to_dict(), ensure_ascii=False)
        except Exception as e:
            return json.dumps({"passed": False, "reason": f"Error: {e}"}, ensure_ascii=False)

    def to_code_prompt(self) -> str:
        return (
            "def checkexcelfile(file_path: str, things_to_check: str) -> str:\n"
            "    '''Verify an Excel (.xlsx) file according to a simple checks string.\n"
            "    Note: file_path must be a Host path; if the file is in the VM, first use get_vm_file to download it.\n"
            "    Returns a JSON string with keys passed, reason, llm_code_path, stdout, stderr.'''\n"
        )
